import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
from collections import deque
import random
from collections import defaultdict
from typing import List, Dict
import yaml
import copy
import os
import sys
import torch.nn.functional as F
from torch.optim.lr_scheduler import StepLR
import time
from torch.utils.tensorboard import SummaryWriter
from openpyxl import Workbook,load_workbook
import datetime

from single_equ_SA.section_1 import section_1
from single_equ_SA.section_2 import section_2
from single_equ_SA.section_3 import section_3
from single_equ_SA.section_4 import section_4
from utils import (set_global_seed, log, process_single_data,
                   milp_test ,pso_test,choose_price)
from price_retrieval import PriceQuery
os.chdir(sys.path[0])


class Critic(nn.Module):
    def __init__(self,config):
        super().__init__()
        self.state_dim = config['state_dim'] 
        self.action_dim = config['action_dim'] 

        self.joint_net = nn.Sequential(
            nn.Linear(self.state_dim + self.action_dim, 128),
            nn.LeakyReLU(0.1),
            nn.Linear(128, 256),
            nn.LeakyReLU(0.1),
            nn.Linear(256, 512),
            nn.LeakyReLU(0.1),
            nn.Linear(512, 256),
            nn.LeakyReLU(0.1),
            nn.Linear(256, 128),
            nn.LeakyReLU(0.1),
            nn.Linear(128, self.action_dim),
        )

    def forward(self, state, action):
        sa = torch.cat([state, action], -1)
        return self.joint_net(sa)


class Actor(nn.Module):
    def __init__(self,config):
        super().__init__()
        self.state_dim = config['state_dim'] 
        self.action_dim = config['action_dim'] 
        self.device = config['device'] 

        self.base_net = nn.Sequential(
            nn.Linear(self.state_dim, 128),
            nn.LeakyReLU(0.1),
            nn.Linear(128, 256),
            nn.LeakyReLU(0.1),
            nn.Linear(256, 128),
            nn.LeakyReLU(0.1),
            nn.Linear(128, 64),
            nn.LeakyReLU(0.1),
            nn.Linear(64, self.action_dim),
        )

    def forward(self, state , last_action=None):
        x = torch.sigmoid(self.base_net(state))
        return x   

class TD3:
    def __init__(self, config):
        super().__init__()
        self.config = config
        self.device = config["device"]
        device = config["device"]
        self.Q1 = Critic(config)
        self.actor = Actor(config)

        self.Q1 = self.Q1.to(device)
        self.actor = self.actor.to(device)

    def get_action(self, state , last_action):
        action = self.actor(state , last_action)
        return action


    def compute_value_loss(self, config, s_batch, a_batch, next_s_batch, r_batch, d_batch):
        with torch.no_grad():
            a = self.actor(next_s_batch , a_batch)
            device = config['device']
            
            noise = torch.clamp(
                torch.randn_like(a) * config['actor_noise']["policy_noise"],
                -config['actor_noise']["noise_clip"],
                config['actor_noise']["noise_clip"],
            )
            
            min_action = a_batch - config['smoothing_factor']
            max_action = a_batch + config['smoothing_factor']
            base_0 = torch.tensor([[0]],dtype = torch.float32).to(device)
            base_1 = torch.tensor([[1]],dtype = torch.float32).to(device) 
            a = torch.clamp(a + noise, min= torch.max(base_0, min_action) , max= torch.min(base_1, max_action))

            assert abs(a - a_batch).max().item() <= config['smoothing_factor'] + 1e-6 , f"Smooth constraint failed"

            q1 = self.Q1(next_s_batch, a).squeeze() 

            # Compute TD target
            y = r_batch + config["gamma"] * q1 * (1 - d_batch)

        qvals1 = self.Q1(s_batch, a_batch).squeeze()
        value_loss1 = F.mse_loss(y, qvals1)
        value_loss2 = 0
        return value_loss1, value_loss2

    def compute_policy_loss(self, s_batch):
        a = self.actor(s_batch , s_batch[:,-3:] )
        policy_loss = -self.Q1(s_batch, a).mean()
        return policy_loss



class INFO:
    def __init__(self):
        self.log = defaultdict(list)
        self.episode_length = 0
        self.episode_reward = 0
        self.max_episode_reward = -float("inf")

    def put(self, done, reward):
        if any(done) is True:
            self.episode_length += 1
            self.episode_reward += reward
            self.log["episode_length"].append(self.episode_length)
            self.log["episode_reward"].append(self.episode_reward)

            self.episode_length = 0
            self.episode_reward = 0
        else:
            self.episode_length += 1
            self.episode_reward += reward

class MultiSectionEnv:
    def __init__(self, config):
        self.agent = TD3(config)
        self.info  = INFO()
        self.config = config
        self.device = config["device"]
        device = config["device"]
        self.run_t = 0

        self.replay_pool = [
            torch.empty(0, config["state_dim"]).to(device),  # state
            torch.empty(0, config["action_dim"]).to(device),   # action
            torch.empty(0, config["state_dim"]).to(device),  # next_state
            torch.empty(0, 1).to(device),    # reward
            torch.empty(0, 1).to(device)      # done
        ]

        self.good_replay_pool = [
            torch.empty(0, config["state_dim"]).to(device),  # state
            torch.empty(0, config["action_dim"]).to(device),   # action
            torch.empty(0, config["state_dim"]).to(device),  # next_state
            torch.empty(0, 1).to(device),    # reward
            torch.empty(0, 1).to(device)      # done
        ]

        self.sections = []
        for idx in range(self.config["max_devices"]):
            dev_type = self.config["devices"][idx]["type"]
            if dev_type == "section_1":
                self.sections.append(section_1(self.config, idx))
            elif dev_type == "section_2":
                self.sections.append(section_2(self.config, idx))
            elif dev_type == "section_3":
                self.sections.append(section_3(self.config, idx))
            elif dev_type == "section_4":
                self.sections.append(section_4(self.config, idx))

        # Initialize global state
        self.buffers = []
        self.product = 0.0
        self.max_hours = 24  
        self.episode_reward = 0

        self.Q1_optimizer = optim.Adam(self.agent.Q1.parameters(), lr=config["critic"]["lr"])   
        self.actor_optimizer = torch.optim.Adam(self.agent.actor.parameters(), lr=config["actor"]["lr"])

        self.Q1_scheduler = StepLR(self.Q1_optimizer, step_size=500, gamma=0.5)

        self.writer = SummaryWriter(log_dir="/root/tf-logs/")
    
    def reset(self,is_train, real_price , DAH_price , last_average_price):
        if is_train:
            self.initial_buffers = []
            for idx in range(self.config["max_devices"]):
                self.initial_buffers.append(np.random.uniform(0, 0.4))
                #Initialize only buffer3, see iPad. 0.4 to 0.8 just ensures that J is above 5. And 4 is exactly the case of J=6, and 0.37 is the case of J=5.5
            self.initial_buffers[-1] = np.random.uniform(0.6, 0.7) 

            self.buffers = self.initial_buffers

            self.product = 0.0

            query = PriceQuery()  
            self.train_price, lase_24_average = query.get_24h_train()

            max_price = max(self.train_price[:24])
            min_price = min(self.train_price[:24])
            self.sections[2].normalizated_price(max_price, min_price)
            self.sections[3].normalizated_price(max_price, min_price)

            self.averge_price = sum(self.train_price[:24])/24

            return 0
        if not is_train:
            self.initial_buffers = [0.3,0.3,0.3,0.7]
            self.buffers = self.initial_buffers


            max_price = max(real_price[:24])
            min_price = min(real_price[:24])
            self.sections[2].normalizated_price(max_price, min_price)
            self.sections[3].normalizated_price(max_price, min_price)


            self.train_price = real_price
            self.averge_price = sum(DAH_price[:24])/24

            return 0
    def train_one_episode(self, is_train, training_step , real_price=[],DAH_price=[] , last_average_price = None, filename='',Test=False):
        """训练一个完整周期"""
        self.reset(is_train, real_price , DAH_price, last_average_price)   #state is a dictionary, i.e. state['buffers'] state['products'] state['hour']
        device = self.device
        self.buffer_24h = {
            "s": defaultdict(dict),
            "a":defaultdict(dict),
            "ns":defaultdict(dict),
            "r": defaultdict(float),
            "done": defaultdict(float) 
        }
        self.episode_reward = []
        test_action = [[],[],[],[]]

        current_price = self.train_price[0]

        done_list = [False, False, False, False]
        done_24h = False



        _, milp_real_real ,x_milp = milp_test(self.train_price[:24], self.train_price[:24], config ,"test_ele.xlsx",self.initial_buffers,save = False)

        for time in range(self.max_hours):

            self.run_t = time

            all_reward = torch.tensor([[0]],dtype=torch.float32).to(device)
           
           # 1-Definition of state
            if time == 0 :
                b1_ratio = torch.tensor([[self.buffers[0]]],dtype=torch.float32).to(device)
                b2_ratio = torch.tensor([[self.buffers[1]]],dtype=torch.float32).to(device)
                b3_ratio = torch.tensor([[self.buffers[2]]],dtype=torch.float32).to(device)
                b4_ratio = torch.tensor([[self.buffers[3]]],dtype=torch.float32).to(device)
                
                price = torch.tensor([[current_price / self.averge_price]],dtype=torch.float32).to(device)

                remaining_t = torch.tensor([[(24-(time%24)) / 24]],dtype=torch.float32).to(device)

                s = torch.cat([b1_ratio.clone(), b2_ratio.clone(), b3_ratio.clone(), 
                               b4_ratio.clone(), price.clone(), remaining_t.clone(), 
                               ], dim=1)
            else:
                s =  ns

            if time % 24 not in self.buffer_24h["s"]:
                self.buffer_24h["s"][time % 24] = {}
            self.buffer_24h["s"][time%24] = s    

           # 2-Determination of action  
            if is_train: 
                if training_step < self.config["warmup_steps"]:  # warmup_stepsFirst take 10 and then make it bigger when it runs.
                    action = []
                    a1 = x_milp[time][0]
                    action.append(a1)
                    a3 = x_milp[time][1]
                    action.append(a3)
                    a4 = x_milp[time][2]
                    action.append(a4)
                    
                    assert all(-1e-6 <= a <= 1+1e-6 for a in action), f"Action values ​​do not have to be in the range [0,1]."

                else:
                    action = self.agent.get_action(s , s[: , -3:])
                    
                    action = action.cpu().data.numpy()

                    initial_noise_std = self.config["actor_noise"]["initial_noise_std"]        # Initial noise standard deviation (50% of range)
                    noise_decay = self.config["actor_noise"]["noise_decay"]                    # Decay coefficient every xx rounds
                    min_noise_std = self.config["actor_noise"]["min_noise_std"]                       # Minimum noise standard deviation (10% of range)
                    decay_interval = self.config["actor_noise"]["decay_interval"]                       # Decay interval (rounds)

                    # Calculate the current decay stage
                    decay_steps = (training_step - self.config["warmup_steps"]) // decay_interval  
                    # Calculate the current noise standard deviation (based on training rounds)
                    noise_std = initial_noise_std * (noise_decay ** decay_steps)
                    noise_std = max(noise_std, min_noise_std)  # Avoid too little noise

                    action_noise = np.random.randn(self.config['action_dim']) * noise_std
                    action_noise = torch.tensor(action+action_noise, dtype=torch.float32).unsqueeze(0)

                    a_batch = s[:, -3:]
                    # min_action = a_batch - self.config['smoothing_factor']
                    # max_action = a_batch + self.config['smoothing_factor']
                    base_0 = torch.tensor([0],dtype = torch.float32)
                    base_1 = torch.tensor([1],dtype = torch.float32) 
                    a = torch.clamp(action_noise, min= base_0 , max=base_1)


                    action = a.data.numpy().flatten().tolist()
                    assert all(-1e-6 <= x <= 1+1e-6 for x in action), f"Action values ​​do not have to be in the range [0,1]"
            
            else: 
                action = self.agent.get_action(s,s[: , -3:]).squeeze().tolist()
                assert all(-1e-6 <= x <= 1+1e-6 for x in action), f"Action values ​​do not have to be in the range [0,1]"

            if time % 24 not in self.buffer_24h["a"]:
                self.buffer_24h["a"][time % 24] = {}
            tensor_action = torch.tensor(action, dtype=torch.float32).unsqueeze(0).to(device)
            self.buffer_24h["a"][time%24] = tensor_action    
            assert tensor_action.shape == torch.Size([1, 3]), f"tensor_action tensor shape error"


           # 3-Run on device
            for section_idx, dev in enumerate(self.sections):
                up_buffer_ratio = self.buffers[section_idx - 1]
                own_buffer_ratio = self.buffers[section_idx]
                price = current_price
                product = self.product
                remaining_t = 24-(time%24)
                
                # Step in is all numbers  
                up_buffer_ratio, own_buffer_ratio, product, reward, done_list[section_idx] , section_action, last_action_ratio = dev.step(up_buffer_ratio, own_buffer_ratio, 
                                                                                                                                            price, product, remaining_t, Test, 
                                                                                                                                            training_step, action, x_milp[time])      
                
                test_action[section_idx].append(section_action)
                assert up_buffer_ratio <= 2.5, f"The up_buffer_ratio value is abnormal"
                assert own_buffer_ratio <= 2.5, f"The up_buffer_ratio value is abnormal"

                all_reward +=reward
                assert all_reward <= 6 + 1e-6, f"all_reward value is abnormal"
                self.product = product

                self.buffers[section_idx - 1] = up_buffer_ratio
                self.buffers[section_idx] = own_buffer_ratio

           #4-Determine ns data   
            current_price = self.train_price[time + 1]

            b1_ratio = torch.tensor([[self.buffers[0]]],dtype=torch.float32).to(device)
            b2_ratio = torch.tensor([[self.buffers[1]]],dtype=torch.float32).to(device)
            b3_ratio = torch.tensor([[self.buffers[2]]],dtype=torch.float32).to(device)
            b4_ratio = torch.tensor([[self.buffers[3]]],dtype=torch.float32).to(device)
            
            price = torch.tensor([[current_price / self.averge_price]],dtype=torch.float32).to(device)

            remaining_t = torch.tensor([[(24-((time + 1)%24))/24]],dtype=torch.float32).to(device)


            ns = torch.cat([b1_ratio.clone(), b2_ratio.clone(), b3_ratio.clone(), 
                            b4_ratio.clone(), price.clone(), remaining_t.clone(), 
                            ], dim=1)

            if not Test:
                if time == (self.max_hours - 1) and not any(done_list):# When the final end is reached without making any mistakes

                    hourly_electric = []
                    for t in range(24):
                        # Extract the parameters of Section 3 and Section 4
                        a3 = test_action[2][t]  
                        a4 = test_action[3][t]  
                        cost_t = (272.61 * a3**2 + 3937.5*a3)*self.train_price[t] + 0.0050219*a4*self.train_price[t]
                        hourly_electric.append(cost_t)
            
                    if sum(hourly_electric) > milp_real_real + 30000:
                        done_list[-1] = True


            if any(done_list):#Whenever death
                all_reward -= 30
            if time == (self.max_hours - 1) and not any(done_list):
                all_reward += 30




            self.episode_reward.append(all_reward)
            if time % 24 not in self.buffer_24h["r"]:
                self.buffer_24h["ns"][time % 24] = {}
                self.buffer_24h["r"][time % 24] = {}  
            self.buffer_24h["ns"][time%24] = ns    
            self.buffer_24h["r"][time%24] = all_reward


           # Interception at 24h
            if Test:
                if time == (self.max_hours - 1):
                    if  not any(done_list): 
                        self.buffer_24h["done"][time%24] = torch.tensor([[any(done_list)]], dtype=torch.float32).to(device) 
                        test_action = self.test_ppc(test_action)
                        total_cost =  self.test_calculate(test_action , real_price, filename)
                        return total_cost
                    else:
                        raise ValueError("He should have died at 24 hours")

            #6-Subsequent processing  
            if time == (self.max_hours - 1):
                done_24h = True
                        
            self.update_product(time, is_train,done_list)                
            if any(done_list) or done_24h:
                if self.buffer_24h["s"] or self.buffer_24h["r"]:  
                    self.buffer_24h["done"][time%24] = torch.tensor([[any(done_list)]], dtype=torch.float32).to(device)
                    self._store_cache(is_train)

                if Test:
                    raise ValueError("Died before 24 hours")


                if not is_train:
                    episode_reward_sum = 0
                    for t in range(len(self.episode_reward)):
                        episode_reward_sum +=  self.episode_reward[t]
                    log(self.writer, episode_reward_sum, "test_reward", training_step - config["warmup_steps"] )
                    
                    if any(done_list) : # It means that there is something wrong with
                        log(self.writer, 0, "test_run", training_step - config["warmup_steps"] )
                        return 2e6
                    else :
                        hourly_electric = []
                        for t in range(24):
                            # Extract the parameters of Section 3 and Section 4
                            a3 = test_action[2][t]  
                            a4 = test_action[3][t]  
                            cost_t = (272.61 * a3**2 + 3937.5*a3)*real_price[t] + 0.0050219*a4*real_price[t]
                            hourly_electric.append(cost_t)

                        log(self.writer, 1, "test_run", training_step - config["warmup_steps"] )

                        return sum(hourly_electric)

                if is_train and training_step >= self.config["warmup_steps"]:
                    if any(done_list) : #
                        log(self.writer, 0, "train_run", training_step - config["warmup_steps"] )
                        print(0)
                    if not any(done_list) and done_24h:# done_24 is guaranteed to be at the 24th hour, the former guarantees that there is no problem at the 24th hour and
                        log(self.writer, 1, "train_run", training_step - config["warmup_steps"] )
                        print(1)
                return None

            self.buffer_24h["done"][time%24] = torch.tensor([[any(done_list)]], dtype=torch.float32).to(device)

    def update_product(self, t, is_train,done_list):
        device = self.config['device']
        if t % 24  == 23:# it is triggered at the 24th hour
            self.product =0 
            self.buffer_24h["done"][t%24] = torch.tensor([[any(done_list)]], dtype=torch.float32).to(device)
            self._store_cache(is_train)
            
            return 0

    def _store_cache(self , is_train):
        """ Deep copy the current cache data """
        import copy
        if is_train:
            cache_data = {
                "s": copy.deepcopy(self.buffer_24h["s"]),
                "a": copy.deepcopy(self.buffer_24h["a"]),
                "ns": copy.deepcopy(self.buffer_24h["ns"]),
                "r": copy.deepcopy(self.buffer_24h["r"]),
                "done": copy.deepcopy(self.buffer_24h["done"]) 
            }
            
            s, a, ns, r, done = process_single_data(cache_data) #(time,7) (time,3) (time,7) (time,1) (time,1)
            
            
            if training_step >= self.config["warmup_steps"]:
                self.replay_pool = [
                    torch.cat([self.replay_pool[0], s], dim=0),
                    torch.cat([self.replay_pool[1], a], dim=0),
                    torch.cat([self.replay_pool[2], ns], dim=0),
                    torch.cat([self.replay_pool[3], r], dim=0),
                    torch.cat([self.replay_pool[4], done], dim=0)
                ]
                current_length = self.replay_pool[0].size(0)
                if current_length > self.config["memory_capacity"]:
                    keep_from = current_length - self.config["memory_capacity"]
                    self.replay_pool = [t[keep_from:] for t in self.replay_pool]

            else:
                self.good_replay_pool= [
                    torch.cat([self.good_replay_pool[0], s], dim=0),
                    torch.cat([self.good_replay_pool[1], a], dim=0),
                    torch.cat([self.good_replay_pool[2], ns], dim=0),
                    torch.cat([self.good_replay_pool[3], r], dim=0),
                    torch.cat([self.good_replay_pool[4], done], dim=0)
                ]

            # Clear the cache but keep the data structure
            self.buffer_24h["s"].clear()
            self.buffer_24h["a"].clear()
            self.buffer_24h["ns"].clear()
            self.buffer_24h["r"].clear()
            self.buffer_24h["done"].clear()
            
            
            return 0
        else : 
            return 0
        
    def test_ppc(self, test_action):
        target_pro = self.config["target_pro"]
        max_action = self.config["devices"][2]["max_x"]

        total = 7.75 * sum(test_action[2])    
	    # The output has reached the target and is returned directly
        if total >= target_pro:
            return test_action
        else:
            deficit = (target_pro - total) / 7.75
            adjusted_action = test_action.copy()

            # Generates a priority queue with adjustable time periods (highest throughput first)
            sorted_indices = sorted(
            range(len(test_action[2])),
            key=lambda i: test_action[2][i],
            reverse=True
            )
        
        intensity = self.config['action_smoothing']  

        # Calculate the single adjustment range based on the adjustment intensity
        for idx in sorted_indices:
            if deficit <= 0:
                break

            current = adjusted_action[2][idx]
            available_space = max_action - current
            
            if available_space <= 0:
                continue
                
            adjust_amount = min(
                deficit,  
                available_space * (1 - intensity * 0.5)  
            )
            
            adjusted_action[2][idx] += adjust_amount
            deficit -= adjust_amount

        return adjusted_action
    def test_calculate(self, test_action , ele_price, filename):        
        hourly_electric = []
        for t in range(24):
            # Extract the parameters of Section 3 and Section 4
            a3 = test_action[2][t]  
            a4 = test_action[3][t]  
            cost_t = (272.61 * a3**2 + 3937.5*a3)*ele_price[t] + 0.0050219*a4*ele_price[t]
            hourly_electric.append(cost_t)
    
        try:
            wb = load_workbook(filename)
            if 'TD3_Results' in wb.sheetnames:
                del wb['TD3_Results']  
        except :
            return sum(hourly_electric)

        ws = wb.create_sheet('TD3_Results', 3)  

        max_a3 = max(test_action[2])
        min_a3 = min(test_action[2])
        ac_normal = []
        for i in range(24):
            ac_normal.append((test_action[2][i] - min_a3)/(max_a3 - min_a3))

        data = [
        [''] + list(range(1,25)),
        ['td3-price'] + [c for c in hourly_electric],
        ['td3-total', sum(hourly_electric)],
        ['action1'] + [a for a in test_action[0]],
        ['action2'] + [a for a in test_action[1]],
        ['action3'] + [a for a in test_action[2]],
        ['action4'] + [a for a in test_action[3]],
        ['action3_normal'] + [a for a in ac_normal],
        ['product', sum(test_action[2])*7.75],
    ]
        for row in data:
            ws.append(row)
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=2):
            for cell in row:
                cell.number_format = '#,##0.00'

        for row in ws.iter_rows(min_row=4, max_row=7, min_col=2):
            for cell in row:
                cell.number_format = '0.000'
        wb.save(filename)
        print("CORE-DDPG test completed! Results saved")

    def learn(self, training_step,is_train):
        config = self.config
        # if self.replay_pool[0].size(0) < config["start_capacity"]:
        if training_step < config["warmup_steps"]:
            return training_step
        else:
            bs, ba, bns, br, bdone = self.get_batch_data(config)

            #Step 2: Update Q1 and Q2
            value_loss1, value_loss2 = self.agent.compute_value_loss(config, bs, ba, bns, br, bdone)
                                                                     
            self.Q1_optimizer.zero_grad()
            value_loss1.backward()
            self.Q1_optimizer.step()

            log(self.writer, value_loss1, "critic", training_step - config["warmup_steps"] )

            episode_reward_sum = 0
            for t in range(len(self.episode_reward)):
                episode_reward_sum +=  self.episode_reward[t]
            log(self.writer, episode_reward_sum, "episode_reward", training_step - config["warmup_steps"])

            #Step 3: Update the actor
            if training_step % config["K"] == 0 and training_step!=0:

                policy_loss = self.agent.compute_policy_loss(bs)
                self.actor_optimizer.zero_grad()
                policy_loss.backward()
                self.actor_optimizer.step()

                #log：actor——loss
                log(self.writer, policy_loss, "actor", training_step-config["warmup_steps"])
  
            log(self.writer, self.run_t, "run_time", training_step - config["warmup_steps"])

            return None

    def get_batch_data(self, config):
        bsz_bad = self.replay_pool[0].size(0) if self.replay_pool[0].size(0) < int(config["bsz"] * 3 / 4) else int(config["bsz"] * 3 / 4)
        indices = torch.randperm(self.replay_pool[0].size(0))[:bsz_bad]  # Sampling without replacement  
        bs_bad = self.replay_pool[0][indices]
        ba_bad = self.replay_pool[1][indices]
        bns_bad = self.replay_pool[2][indices]
        br_bad = self.replay_pool[3][indices]
        bdone_bad = self.replay_pool[4][indices]
            
        bsz_good = self.good_replay_pool[0].size(0) if self.good_replay_pool[0].size(0) < int(config["bsz"] * 1 / 4) else int(config["bsz"] * 1 / 4)
        indices = torch.randperm(self.good_replay_pool[0].size(0))[:bsz_good]  # Sampling without replacement  
        bs_good = self.good_replay_pool[0][indices]
        ba_good = self.good_replay_pool[1][indices]
        bns_good = self.good_replay_pool[2][indices]
        br_good = self.good_replay_pool[3][indices]
        bdone_good = self.good_replay_pool[4][indices]
        #Sequential Merge
        bs_combined = torch.cat([bs_bad, bs_good], dim=0)
        ba_combined = torch.cat([ba_bad, ba_good], dim=0)
        bns_combined = torch.cat([bns_bad, bns_good], dim=0)
        br_combined = torch.cat([br_bad, br_good], dim=0)
        bdone_combined = torch.cat([bdone_bad, bdone_good], dim=0)

            # Generate random permutation index (shuffle order)
        batch_size = bs_combined.size(0)
        shuffle_indices = torch.randperm(batch_size)

            # Apply the same random permutation (keep data aligned)
        bs = bs_combined[shuffle_indices]
        ba = ba_combined[shuffle_indices]
        bns = bns_combined[shuffle_indices]
        br = br_combined[shuffle_indices]
        bdone = bdone_combined[shuffle_indices]
        return bs,ba,bns,br,bdone


    def save_checkpoint(self, training_step, folder="checkpoints-" + time.strftime("%m-%d-%H:%M")):
        if not os.path.exists(folder):
            os.makedirs(folder)
            
        checkpoint_path = os.path.join(folder, f"checkpoint_{training_step}.pth")
        
        checkpoint = {
            'Q1': self.agent.Q1.state_dict(),
            'actor': self.agent.actor.state_dict(),
            'Q1_optimizer': self.Q1_optimizer.state_dict(),
            'actor_optimizer': self.actor_optimizer.state_dict(),
            'replay_pool': self.replay_pool,
            'training_step': training_step,
            'random_states': {
                'numpy': np.random.get_state(),
                'torch': torch.get_rng_state(),
                'python': random.getstate()
            }
        }
        torch.save(checkpoint, checkpoint_path)
        print(f"Saved checkpoint at step {training_step}")


    def load_checkpoint(self, checkpoint_path):
        checkpoint = torch.load(checkpoint_path)
        
        # Load network parameters
        self.agent.Q1.load_state_dict(checkpoint['Q1'])
        self.agent.actor.load_state_dict(checkpoint['actor'])
        self.Q1_optimizer.load_state_dict(checkpoint['Q1_optimizer'])
        self.actor_optimizer.load_state_dict(checkpoint['actor_optimizer'])
        self.replay_pool = checkpoint['replay_pool']
        np.random.set_state(checkpoint['random_states']['numpy'])
        torch.set_rng_state(checkpoint['random_states']['torch'])
        random.setstate(checkpoint['random_states']['python'])
        
        return checkpoint['training_step'] 



if __name__ == "__main__":
    config_dir = "single_equ_SA/equipment_config.yaml"
    config = yaml.load(open(config_dir, "r", encoding='utf-8'), Loader=yaml.FullLoader)#Use the predefined config file and load the definitions into the config
    
    SEED = config['seed']
    set_global_seed(SEED)
    env = MultiSectionEnv(config)
    TRAIN = True
    training_step = 0
    epoch = 3000000

    # Try loading the latest checkpoint
    use_checkpoints = False
    if use_checkpoints:
        # checkpoint_folder = input("Enter the folder name:")
        checkpoint_folder = "wzytest"
        latest_checkpoint = None
        if os.path.exists(checkpoint_folder):
            checkpoints = [f for f in os.listdir(checkpoint_folder) if f.startswith("checkpoint_")]
            if checkpoints:
                # Find the maximum number of training steps
                latest_step = max([int(f.split("_")[1].split(".")[0]) for f in checkpoints])
                latest_checkpoint = os.path.join(checkpoint_folder, f"checkpoint_{latest_step}.pth")
        
        if latest_checkpoint:
            print(f"Loading checkpoint from {latest_checkpoint}")
            training_step = env.load_checkpoint(latest_checkpoint)
        else:
            training_step = 0
            print("No checkpoint found, starting from scratch")


    for idx in range(3):
        filename = f"{training_step}" +f"idx-{idx}" +  ".xlsx"
        real_price, DAH_price , last_average_price = choose_price(idx , filename,  save = not TRAIN )
        #When training, of course it is not saved, it is just used to get a value; it will be automatically saved when predicting
        buffers_ratios = [0.3,0.3,0.3,0.7]
        milp_DAHP_DAHP, milp_DAHP_real , _ = milp_test(DAH_price, real_price, config ,filename,buffers_ratios,save=not TRAIN)
        pso_DAHP_DAHP, pso_DAHP_real = pso_test(DAH_price, real_price, config ,filename, buffers_ratios ,save=not TRAIN)

        best_grade = 2e6
        # When test=false + train=True, it is training; when test=false + train=false, it is validation set; when test=true + train=false, it is testing
        if TRAIN:
            for episode in range(epoch):# The stage of formally starting to train the network
                if training_step >= 210000:
                    break   
                if training_step >= config['warmup_steps']:
                    cost_td3 = env.train_one_episode(False, training_step,real_price, DAH_price, last_average_price)
                    # The fourth one halfway through is 2e6, so it’s normal that you didn’t die!
                    log(env.writer, (pso_DAHP_real - cost_td3 )/1e4, "cost_difference", training_step - config["warmup_steps"])
                    if cost_td3 < pso_DAHP_real and  cost_td3 < best_grade - 3000:
                        best_grade = cost_td3
                        env.save_checkpoint(training_step , folder="best-checkpoints")

                env.train_one_episode(TRAIN, training_step)
                env.learn(training_step,TRAIN)

                training_step += 1
                if training_step > config['warmup_steps']:
                    if (training_step - config['warmup_steps']) % 10000 == 0 and (training_step - config['warmup_steps']) != 0:
                        env.save_checkpoint(training_step)
        
        if not TRAIN:
            a = env.train_one_episode(TRAIN, training_step,real_price, DAH_price, last_average_price,filename, Test = True)
        print("OK")

        