import yaml
import torch
import numpy as np
import random
import gurobipy as gp
from gurobipy import GRB
import numpy as np
import time
import csv
from datetime import datetime
from openpyxl import Workbook,load_workbook
import copy
from pyswarm import pso
import os,sys
import pandas as pd

def process_single_data(data_entry):
    """Processing a single data entry (data_entry is an element in a list)"""
    time_steps = len(data_entry["s"])
    
    s_data =  []
    a_data =  []
    ns_data =  []
    r_data = []
    done_data = []

    for t in range(time_steps):
        s_data.append(data_entry["s"][t])
        a_data.append(data_entry["a"][t])
        ns_data.append(data_entry["ns"][t])
        r_data.append(data_entry["r"][t])
        done_data.append(data_entry["done"][t])

    return (
        torch.cat(s_data, dim=0),# (time,7)
        torch.cat(a_data, dim=0), # (time,3)
        torch.cat(ns_data, dim=0),# (time,7)
        torch.cat(r_data, dim=0),# (time,1)
        torch.cat(done_data, dim=0)# (time,1)
    )
 
def set_global_seed(seed):
    torch.manual_seed(seed)          # PyTorch
    torch.cuda.manual_seed_all(seed) # CUDA
    np.random.seed(seed)             # NumPy
    random.seed(seed)               # Python
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

def log(writer, value, str, training_step):
    if str == "critic":
        writer.add_scalar("SARL/Critic", value, training_step)
    if str == "actor":
        writer.add_scalar("SARL/actor", value, training_step)
    if str == "episode_reward":
        writer.add_scalar("SARL/episode_reward", value, training_step)
    if str == "train_run":
        writer.add_scalar("SARL/train_run", value, training_step)
    if str == "run_time":
        writer.add_scalar("SARL/run_time", value, training_step) 
    if str == "test_reward":
        writer.add_scalar("SARL/test_reward", value, training_step)
    if str == "test_run":
        writer.add_scalar("SARL/test_run", value, training_step)
    if str == "cost_difference":
        writer.add_scalar("SARL/cost_difference", value, training_step)

def milp_test(DAH_price, real_price, config , filename, buffers_ratio,save=False):
    T = 24  # time range
    sections = [1, 3, 4]#No operation sec_2
    buffers = [1, 2, 3, 4]

    scale_factor = 0.02
    prod_limits = {
    1: (0 + scale_factor , 1 - scale_factor ),
    3: (0 + scale_factor , 1 - scale_factor ),
    4: (0 + scale_factor , 1 - scale_factor )
}

    buffer_limits = {
    1: (0 + scale_factor , 1 - scale_factor ),
    2: (0 + scale_factor , 1 - scale_factor ),
    3: (0 + scale_factor , 1 - scale_factor ),
    4: (0 + scale_factor , 1 - scale_factor )
}

# Sample electricity price data (replace with actual value)
    total_target = 938.5  # target

# ====================
# Model building
# ====================
    model = gp.Model("Production_Scheduling")
    model.Params.OutputFlag = 0  
    model.Params.NonConvex = 2  
# Decision variables
    x = model.addVars(sections, range(T), name="x")  # Output of each section
    b = model.addVars(buffers, range(T+1), name="b")   # Buffer inventory

# ====================
# Constraints
# ====================
# Material balance constraints
    for t in range(T+1):
        if t == 0 :
            model.addConstr(b[1, t] == buffers_ratio[0] )
            model.addConstr(b[2, t] == buffers_ratio[1] )
            model.addConstr(b[3, t] == buffers_ratio[2] )
            model.addConstr(b[4, t] == buffers_ratio[3] )

        else:
            a1 = 90.567 + x[1, t-1] * (120.756 - 90.567)
            a2 = 143.8652073732719
            a3 = 4.5 + x[3, t-1] * (6 - 4.5)
            a4 = 90.567 + x[4, t-1] * (120.756 - 90.567)

            model.addConstr(b[1, t] == (b[1, t-1] * 419.58 + 27.972*(a1/20.126) - a2) / 419.58)
            model.addConstr(b[1, t] >= buffer_limits[1][0])
            model.addConstr(b[1, t] <= buffer_limits[1][1])

            model.addConstr(b[2, t] == (b[2, t-1] * 359.663 + a2 - a3*27.972) / 359.663)
            model.addConstr(b[2, t] >= buffer_limits[2][0])
            model.addConstr(b[2, t] <= buffer_limits[2][1])

            model.addConstr(b[3, t] == (b[3, t-1] * 301.89 + a3*20.126 - a4) / 301.89)
            model.addConstr(b[3, t] >= buffer_limits[3][0])
            model.addConstr(b[3, t] <= buffer_limits[3][1])

            model.addConstr(b[4, t] == (b[4, t-1] * 301.89 + a4 - a1) / 301.89)
            model.addConstr((b[4, t-1] * 301.89 - a1) / 301.89 >= buffer_limits[4][0])#注意这里
            model.addConstr((b[4, t-1] * 301.89 - a1) / 301.89 <= buffer_limits[4][1])



    for i in sections:#i is taken from [1,2,3,4]
        for t in range(T):
            model.addConstr(x[i, t] >= prod_limits[i][0])
            model.addConstr(x[i, t] <= prod_limits[i][1])
 
    # New smoothing constraint: Constraint on the rate of change of production in adjacent time periods (no more than 50% capacity difference)
    for i in sections:
        delta_i = config['smoothing_factor'] * (prod_limits[i][1] - prod_limits[i][0])
        for t in range(T-1):  #t is from 0 to T-2 (inclusive)
            model.addConstr(x[i, t+1] - x[i, t] <= delta_i, name=f"rate_up_{i}_{t}")
            model.addConstr(x[i, t] - x[i, t+1] <= delta_i, name=f"rate_down_{i}_{t}")

    # Total output target (assuming that Section 4 is the final output)
    product = 0
    for t in range(T):
        a3 = 4.5 + x[3, t] * (6 - 4.5)
        product += a3 * 7.75
    
    model.addConstr(product >= total_target)
    model.addConstr(product <= total_target + 46.875)

# ====================
# Objective Function
# ====================
    electric_cost = 0
    for t in range(T):
        a3 = 4.5 + x[3, t] * (6 - 4.5)
        a4 = 90.567 + x[4, t] * (120.756 - 90.567)
        electric_cost += (272.61 * a3 * a3 + 3937.5 * a3)* DAH_price[t] + 5.0219E-3 * a4 * DAH_price[t]

    model.setObjective(electric_cost, GRB.MINIMIZE)


# ====================
# Model Solution
# ====================
    model.optimize()
    if model.status != GRB.OPTIMAL:
        raise ValueError("Failed to solve the model.")
	# ====================
    # Result analysis & export to Excel
    # ====================
    
    if save:
        try:
            wb = load_workbook(filename)
            if 'PSO Results' in wb.sheetnames:
                del wb['PSO Results']  
        except FileNotFoundError:
            wb = Workbook()
            del wb['Sheet'] 
        # Create a new worksheet
        ws = wb.create_sheet('Milp_Results', 1)  # Insert at the second position
        real_hourly_prices = []
        for t in range(T):
            x3_val = 4.5 + x[3,t].x * (6 - 4.5)  # Production of Section 3
            x4_val = 90.567 + x[4,t].x * (120.756 - 90.567) # Production of Section 4
            price_t = (272.61 * x3_val* x3_val + 3937.5 * x3_val) * real_price[t] + 5.0219E-3 * x4_val * real_price[t]
            
            real_hourly_prices.append(price_t)

        matrix  = [
        [''] + list(range(1,26)),
        ['milp-price(¥)'] + [c for c in real_hourly_prices],
        ['milp-total(¥)'] + [sum(real_hourly_prices)],
        ]
        for row in matrix:
            ws.append(row)

        for i in [1,3,4]:  # Traverse 4 sections
            x_values = [f"x{i}"]  # Row header (e.g. "x1")
            for t in range(T):
                x_values.append(x[i, t].x)  
            ws.append(x_values)

        # Write buffer lines (starting from line 8)
        for k in [1,2,3,4]:  
            b_values = [f"b{k}"]  
            for t in range(T+1):  
                b_values.append(b[k, t].x)
            ws.append(b_values)

        # Write action3_normalied lines (starting from line 9)
        action_normal = ["action3-normal"]
        action3 = []
        for t in range(T):
            action3.append(x[3, t].x) 
        max_action = max(action3)
        min_action = min(action3)
        for t in range(T):  # Note that the buffer has T+1 time points (including the initial time)
            action_normal.append((action3[t] - min_action)/(max_action - min_action))
        ws.append(action_normal)

        for row in ws.iter_rows(min_row=2, max_row=2, min_col=2):
            for cell in row:
                cell.number_format = '#,##0.00'

        for row in ws.iter_rows(min_row=4, max_row=7, min_col=2):
            for cell in row:
                cell.number_format = '0.000'

        wb.save(filename)
        print(f"Results saved to:{filename}")
        print("Milp optimization completed! Results saved")

        action = []
        for t in range(T):
            action_t = []
            x1_val = x[1,t].x
            x3_val = x[3,t].x  
            x4_val = x[4,t].x 
            action_t.append(x1_val)
            action_t.append(x3_val)
            action_t.append(x4_val)
            action.append(action_t)

        return round(model.objVal, 2) , sum(real_hourly_prices) , action
    else:

        real_hourly_prices = []
        for t in range(T):
            x3_val = 4.5 + x[3,t].x * (6 - 4.5)  
            x4_val = 90.567 + x[4,t].x * (120.756 - 90.567) 
            price_t = (272.61 * x3_val* x3_val + 3937.5 * x3_val) * real_price[t] + 5.0219E-3 * x4_val * real_price[t]
            
            real_hourly_prices.append(price_t)        
        
        action = []
        for t in range(T):
            action_t = []
            x1_val = x[1,t].x
            x3_val = x[3,t].x  
            x4_val = x[4,t].x 
            action_t.append(x1_val)
            action_t.append(x3_val)
            action_t.append(x4_val)
            action.append(action_t)



        return round(model.objVal, 2) , sum(real_hourly_prices), action

def pso_test(DAH_price, real_price, config, filename, buffers_ratio, save=False):

    sections = [1, 3, 4]
    buffers = [1, 2, 3, 4]

    scale_factor = 0.02
    action_limits = {
    1: (0 + scale_factor , 1 - scale_factor ),
    3: (0 + scale_factor , 1 - scale_factor ),
    4: (0 + scale_factor , 1 - scale_factor )
}

    buffer_limits = {
    1: (0 + scale_factor , 1 - scale_factor ),
    2: (0 + scale_factor , 1 - scale_factor ),
    3: (0 + scale_factor , 1 - scale_factor ),
    4: (0 + scale_factor , 1 - scale_factor )
}

    total_target = 938.5 #1 more margin

    # New: Maximum change allowed in adjacent time periods for each work section
    delta_limits = {
        sec: config['smoothing_factor'] * (action_limits[sec][1] - action_limits[sec][0])
        for sec in sections
    }

    def buffer_simulate(actions):
        """Material tracking with real-time buffer constraints"""
    # 初始化

        buffer = {
                1: buffers_ratio[0],
                2: buffers_ratio[1], 
                3: buffers_ratio[2], 
                4: buffers_ratio[3]
                }
        
        penalty = 0.0
        hourly_violations = [0]*24  
    
        for t in range(24):
            a1_ratio, a3_ratio, a4_ratio = actions[t]  
            a1 = 90.567 + a1_ratio * (120.756 - 90.567)
            a2 = 143.8652073732719
            a3 = 4.5 + a3_ratio * (6 - 4.5)
            a4 = 90.567 + a4_ratio * (120.756 - 90.567)

        # Sec1 -> Sec2
            buffer[1] = (buffer[1] * 419.58 + 27.972*(a1/20.126) - a2) / 419.58

            if buffer[1] < buffer_limits[1][0]:
                violation = buffer_limits[1][0] - buffer[1]
                penalty += 1e6 * violation
                hourly_violations[t] += 1
            elif buffer[1] > buffer_limits[1][1]:
                violation = buffer[1] - buffer_limits[1][1]
                penalty += 1e6 * violation
                hourly_violations[t] += 1
        # Sec2 -> Sec3
            buffer[2] = (buffer[2] * 359.663 + a2 - a3*27.972) / 359.663

            if buffer[2] < buffer_limits[2][0]:
                violation = buffer_limits[2][0] - buffer[2]
                penalty += 1e6 * violation
                hourly_violations[t] += 1
            elif buffer[2] > buffer_limits[2][1]:
                violation = buffer[2] - buffer_limits[2][1]
                penalty += 1e6 * violation
                hourly_violations[t] += 1

        # Sec3 -> Sec4 
            buffer[3] = (buffer[3] * 301.89 + a3*20.126 - a4) / 301.89

            if buffer[3] < buffer_limits[3][0]:
                violation = buffer_limits[3][0] - buffer[3]
                penalty += 1e6 * violation
                hourly_violations[t] += 1
            elif buffer[3] > buffer_limits[3][1]:
                violation = buffer[3] - buffer_limits[3][1]
                penalty += 1e6 * violation
                hourly_violations[t] += 1


        # Sec4 -> Sec1  Prioritize judgment and then update!
            buffer4_initial  =  (buffer[4] * 301.89 - a1) / 301.89

            if buffer4_initial < buffer_limits[4][0]:
                violation = buffer_limits[4][0] - buffer4_initial
                penalty += 1e6 * violation
                hourly_violations[t] += 1
            elif buffer4_initial > buffer_limits[4][1]:
                violation = buffer4_initial - buffer_limits[4][1]
                penalty += 1e6 * violation
                hourly_violations[t] += 1

            buffer[4] = (buffer[4] * 301.89 + a4 - a1) / 301.89

        return penalty, buffer, hourly_violations

    def objective_function(particle):
        """Objective function with real-time constraints"""
    # Decoded particles are 4x24 matrices
        actions = np.array(particle).reshape(3,24).T
    
        # Calculate the penalty for production changes in adjacent periods
        rate_penalty = 0.0
        delta_mapping = [1, 3, 4]

        for sec in [0,1,2]:  
            delta = delta_limits[delta_mapping[sec]]  
            for t in range(23):  
                curr = actions[t, sec]
                next_val = actions[t+1, sec]
                # Calculate the change penalty
                if abs(next_val - curr) > delta:
                    rate_penalty += 1e6 * (abs(next_val - curr) - delta)

        # Electricity bill calculation
        total_cost = 0.0
        hourly_electric_cost = []
        for t in range(24):
            a3 = 4.5 + actions[t][1] * (6 - 4.5)
            a4 = 90.567 + actions[t][2]* (120.756 - 90.567)
            cost_t = (272.61 * a3 * a3 + 3937.5*a3)*DAH_price[t] + 0.0050219*a4*DAH_price[t]
            total_cost += cost_t
            hourly_electric_cost.append(cost_t)
    
    # Total output constraint
        product = 0
        for t in range(24):
            a3 = 4.5 + actions[t][1] * (6 - 4.5)
            product += a3 * 7.75

        prod_penalty = abs(product - total_target) * 1e6
    
    # Real-time buffer constraints
        buffer_penalty, final_buffer, violations = buffer_simulate(actions)
    
        return total_cost * 1e-7 + prod_penalty + buffer_penalty + rate_penalty, (hourly_electric_cost, violations)

# ====================
# Run Optimization
# ====================
    lb = [action_limits[sec][0] for sec in [1,3,4] for _ in range(24)]
    ub = [action_limits[sec][1] for sec in [1,3,4] for _ in range(24)]

    options = {
    'swarmsize': 30,
    'omega': 0.4,
    'phip': 0.4,
    'phig': 0.4,
    'maxiter': 200,
    'minstep': 1e-2,
    'minfunc': 1e-2
    }

    best_solution , best_cost = pso(
    lambda x: objective_function(x)[0], 
    lb, ub,
    **options
)

# ====================
# Result processing and output
# ====================
    optimal_actions = np.array(best_solution).reshape(3,24).T
    _, (hourly_electric, violations) = objective_function(best_solution)

    # Verify final yield
    product = 0
    for t in range(24):
        a3 = 4.5 + optimal_actions[t][1] * (6 - 4.5)
        product += a3 * 7.75

    if abs(total_target - product ) > 0.1:
        raise ValueError("警告：产量未达标！")

# Verify hourly_electric and total_cost
    real_total_cost = 0
    real_hourly_cost = []
    action3 = []
    for t in range(24):
        a3 = 4.5 + optimal_actions[t][1] * (6 - 4.5)
        action3.append(a3)
        a4 = 90.567 + optimal_actions[t][2] * (120.756 - 90.567)
        cost_t = (272.61 * a3 * a3 + 3937.5*a3)*real_price[t] + 0.0050219*a4*real_price[t]
        real_hourly_cost.append(cost_t)
        real_total_cost += cost_t
    if abs(real_total_cost - sum(real_hourly_cost)) > 0.1:
        raise ValueError('hourly_price与total_cost不符')
            
    if save:
    # ====================
    # Structuring Excel Data
    # ====================
        try:
            wb = load_workbook(filename)
            if 'PSO Results' in wb.sheetnames:
                del wb['PSO Results']  
        except FileNotFoundError:
            wb = Workbook()
            del wb['Sheet']  
    # Create a new worksheet
        ws = wb.create_sheet('PSO Results', 2) 
        max_a3 = max(action3)
        min_a3 = min(action3)
        action3_normal = [(action3[a] - min_a3)/(max_a3 - min_a3)  for a in range(24)]

        data = [
        [''] + list(range(1,25)),
        ['pso-price'] + [c for c in real_hourly_cost],
        ['pso-total', real_total_cost],
        ['action1'] + [a for a in optimal_actions[:,0]],
        ['action2'] + [143.8652 for a in range(24)],
        ['action3'] + [a for a in optimal_actions[:,1]],
        ['action4'] + [a for a in optimal_actions[:,2]],
        ['action3_normal'] + [a for a in action3_normal],
        ['product', product],
    ]
        for row in data:
            ws.append(row)
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=2):
            for cell in row:
                cell.number_format = '#,##0.00'
 
        for row in ws.iter_rows(min_row=4, max_row=7, min_col=2):
            for cell in row:
                cell.number_format = '0.000'
        wb.save(filename)
        print("PSO optimization completed! Results saved")
        return sum(hourly_electric)  , real_total_cost
    else:
        return  sum(hourly_electric) , real_total_cost

def choose_price(idx, filename, save = False):
    if idx == 0:
        real_price = [
                    1.6, 3.2, 1.7, 2.4, 1.3,
                    0.1, 0.8, 1.8, 2, 1.3,
                    0.8, 1.3, 1.9, 1.5, 1.8,
                    1.7, 1.5, 1.4, 3.1, 6,
                    4.5, 2.9, 2.5, 3.1, 6
                    ]
        
        DAH_price = [
                    1.8, 1.7, 1.8, 2.2, 2.2,
                    2.4, 2.7, 2.9, 2.7, 2.7,
                    2.6, 2.4, 2, 1.7, 1.6,
                    1.7, 1.5, 1.6, 2.1, 4, 
                    4.9, 3.7, 3.7, 2.1,
                    ]
        # 4.17
        last_average_price = 0.8291
    if idx == 1:
        real_price = [
                    2.6, 2.4, 2.3, 2.3, 1.7,
                    1.4, 6.0, 2.4, 1.8, 2.1,
                    2.2, 2.5, 1.3, 1.6, 1.7,
                    2.1, 2.4, 1.7, 2.5, 3.3,
                    8.5, 4.7, 3.3, 4.6, 3.3
                    ]
        
        DAH_price = [
                    2.9, 2.1, 1.9, 1.8, 1.7,
                    2.0, 2.9, 3.7, 3.1, 2.8,
                    2.7, 2.5, 2.3, 2.1, 2.1,
                    2.5, 2.2, 2.5, 4.2, 7.9,
                    8.8, 3.8, 3.1, 2.1, 3.1
                        ]
        # 4.21
        last_average_price = 4.025

    if idx == 2:
        real_price = [
                    2.3, 2.2, 0.9, 0.9, 1.2,
                    -0.7, 5.4, 4.1, 2.7, 2.5,
                    2.6, 8.8, 2.7, 5.5, 2.2,
                    3.2, 17.3, 5.0, 2.3, 5.0,
                    10.2, 7.2, 6.2, 6.0, 6.2
                    ]
        
        DAH_price = [
                    2.7, 2.5, 2.2, 2.1, 2.4,
                    2.5, 4.0, 4.4, 3.4, 3.2,
                    3.0, 3.1, 3.3, 3.4, 3.5,
                    3.7, 4.0, 4.9, 6.7, 9.9,
                    8.8, 5.7, 4.1, 3.0, 4.1
                    ]
        # 4.22
        last_average_price = 2.8083

    if idx == 4:
        real_price = [
                    3.3, 2.6, 2.6, 2.5, 0.3, 
                    -2.8, -1.0, -0.9, -0.9, 0.9,
                    2.0, 1.7 ,1.4 , 1.0, 1.6,
                    1.2, 1.0, 1.5, 7.8, 2.4,
                    4.5, 2.7, 6.3, 3.9, 2.7 
                    ]
        
        DAH_price = [
                    2.3, 2.5, 1.4, 0.8, 0.3,
                    -0.2, 0, -0.8, -0.8, -1.0,
                    -1.0, -1.1,-1.2, -1.1, -1.1,
                    -1.0, -0.7, 0, 0.1, 1.1,
                    1.6, 1.6, 1.2, 0.7, 1.2 
                        ]


    if save:
        wb = Workbook()
        ws = wb.active
        ws.title = "ele_price"

        max_DAH = max(DAH_price)
        min_DAH = min(DAH_price)
        DAH_price_normal = []
        for t in range(len(DAH_price)):
            DAH_price_normal.append((DAH_price[t] - min_DAH) / (max_DAH - min_DAH))

        max_real = max(real_price)
        min_real = min(real_price)
        real_price_normal = []
        for t in range(len(real_price)):
            real_price_normal.append((real_price[t] - min_real) / (max_real - min_real))

        ws.append([''] + list(range(1, 25+1)))

        ws.append(['DAH-price'] + DAH_price)
        ws.append(['real_price'] +  real_price)
        ws.append(['DAH_price_normal'] + DAH_price_normal)
        ws.append(['real_price_normal'] + real_price_normal)
        # Write action3_normalied lines (starting from line 9)

        wb.save(filename)
        print(f"结果已保存至：{filename}")
        print("电价ele结果已保存")
    
    return real_price, DAH_price, last_average_price